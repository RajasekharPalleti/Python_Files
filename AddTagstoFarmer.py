# Author: Rajasekhar Palleti
# This script ensures that if a tag already exists for a farmer, it is retained.
# If a tag is missing, it is added to the master list and assigned to the farmer.
# Duplicate tags are prevented.

import json
import requests
import openpyxl
import time
from GetAuthtoken import get_access_token


def fetch_tags(token, tags_api_url, required_tag):
    """
    Fetch a specific tag from the API and ensure it exists.
    If the required tag is missing, add it via the API.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    try:
        print(f"⏳ Fetching tag '{required_tag}' from the tags API...")
        response = requests.get(tags_api_url, headers=headers)
        response.raise_for_status()
        tags_data = response.json()

        # If API returns empty, manually add the required tag
        if not tags_data:
            print(f"⚠️ No tags found in API. Adding '{required_tag}' manually.")
            return add_tag(token, required_tag)

        # Check if the required tag exists in the fetched data
        for tag in tags_data:
            if tag["name"] == required_tag:
                print(f"✅ Tag '{required_tag}' found in the tags API.")
                return tag

        # If the tag is missing, add it
        return add_tag(token, required_tag)

    except requests.exceptions.RequestException as e:
        print(f"❌ Error fetching tags from API: {e}. Continuing execution.")
        return add_tag(token, required_tag)


def add_tag(token, tag_name):
    """
    Adds a missing tag to the master list via the API.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    print(f"⏳ Adding missing tag: {tag_name}")
    tag_payload = {"name": tag_name, "id": None}

    post_response = requests.post(post_tag_url, headers=headers, json=tag_payload)
    if post_response.status_code == 201:
        print(f"✅ Successfully added missing tag: {tag_name}")
        return post_response.json()
    else:
        print(f"❌ Failed to add tag: {tag_name} - {post_response.text}")
        return None


def post_data_to_api(api_url, token, input_excel, output_excel, tags_api_url):
    """
    Read farmer IDs and required tags from an Excel sheet, fetch existing farmer data,
    append missing tags, and update the farmer data via an API.
    """
    wb = openpyxl.load_workbook(input_excel)
    sheet = wb[sheet_name]

    headers = {"Authorization": f"Bearer {token}"}

    # Add status and failure reason columns if not already present
    if sheet.cell(1, sheet.max_column).value != "Status":
        sheet.cell(1, sheet.max_column + 1, "Status")
    if sheet.cell(1, sheet.max_column).value != "Failure Reason":
        sheet.cell(1, sheet.max_column + 1, "Failure Reason")

    status_col = sheet.max_column - 1
    failure_reason_col = sheet.max_column

    for row in range(2, sheet.max_row + 1):
        farmer_id = sheet.cell(row=row, column=1).value  # Assume Farmer IDs are in column 1
        required_tags = sheet.cell(row=row, column=3).value  # Assume required tags are in column 3

        if required_tags:
            required_tags = [tag.strip() for tag in required_tags.split(",") if tag.strip()]
        else:
            required_tags = []

        print(f"⏳ Processing row {row} - Farmer ID: {farmer_id}, Required Tags: {required_tags}")

        if not farmer_id or not required_tags:
            sheet.cell(row=row, column=status_col, value="=> Skipped: Missing Data")
            continue

        # Fetch or add each required tag
        tags_data = [fetch_tags(token, tags_api_url, tag) for tag in required_tags if tag]
        tags_data = [tag for tag in tags_data if tag is not None]

        if not tags_data:
            sheet.cell(row=row, column=status_col, value="Failed: Tags Not Available")
            continue

        try:
            print(f"⏳ Fetching details for Farmer ID: {farmer_id}")
            get_response = requests.get(f"{api_url}/{farmer_id}", headers=headers)
            get_response.raise_for_status()
            farmer_data = get_response.json()

            # Add the new tags to the farmer
            if "data" in farmer_data and isinstance(farmer_data["data"], dict):
                existing_tags = {tag["name"] for tag in farmer_data["data"].get("tags", [])}
                new_tags = [tag for tag in tags_data if tag["name"] not in existing_tags]

                if new_tags:
                    if "tags" not in farmer_data["data"] or farmer_data["data"]["tags"] is None:
                        farmer_data["data"]["tags"] = []
                    farmer_data["data"]["tags"].extend(new_tags)
                    print(f"✅ Added new tags {new_tags} to Farmer ID: {farmer_id}")
                else:
                    print(f"⚠️ Tags already exist for Farmer ID: {farmer_id}, skipping update.")
                    sheet.cell(row=row, column=status_col, value="Skipped: Tags Already Present")
                    continue
            else:
                print(f"❌ Failed to update tags for Farmer ID: {farmer_id} - No 'data' key found")
                sheet.cell(row=row, column=status_col, value="Failed: No data key in farmer data")
                sheet.cell(row=row, column=failure_reason_col, value=str(farmer_data))
                continue

            time.sleep(0.5)

            # Send updated farmer data to the API
            multipart_data = {"dto": (None, json.dumps(farmer_data), "application/json")}
            print(f"⏳ Posting updated data for Farmer ID: {farmer_id}")
            put_response = requests.put(api_url, headers=headers, files=multipart_data)
            put_response.raise_for_status()
            sheet.cell(row=row, column=status_col, value="✅Success")
            print(f"✅ Successfully posted updates for Farmer ID: {farmer_id}")

        except requests.exceptions.RequestException as e:
            error_message = str(e)
            if hasattr(e, 'response') and e.response is not None:
                error_message = e.response.text
            sheet.cell(row=row, column=status_col, value="❌Failed")
            sheet.cell(row=row, column=failure_reason_col, value=error_message)
            print(f"❌ Failed to update Farmer ID: {farmer_id} - Error: {error_message}")

        time.sleep(0.5)

    wb.save(output_excel)
    print(f"✅ Excel file updated and saved as {output_excel}")


if __name__ == "__main__":
    input_excel = "C:\\Users\\rajasekhar.palleti\\Downloads\\Farmer_tags_gpi.xlsx"
    sheet_name = "Data"
    output_excel = "C:\\Users\\rajasekhar.palleti\\Downloads\\Farmer_tags_gpi_output.xlsx"
    api_url = "https://cloud.cropin.in/services/farm/api/farmers"
    tags_api_url = "https://cloud.cropin.in/services/master/api/tags?size=5000"
    post_tag_url = "https://cloud.cropin.in/services/master/api/tags"
    environment = "prod1"

    print("⏳ Retrieving access token...")
    token = get_access_token("gpi", "9120232024", "Anilkumarkolla", environment)

    if token:
        print("✅ Access token retrieved successfully")
        post_data_to_api(api_url, token, input_excel, output_excel, tags_api_url)
    else:
        print("❌ Failed to retrieve access token. Process terminated.")
