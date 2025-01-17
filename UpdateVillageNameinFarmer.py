import requests
import openpyxl
import time

# Constants
EXCEL_FILE = "C:\\Users\\rajasekhar.palleti\\Downloads\\Pidilite_failed_records.xlsx"
GET_API_URL = "https://cloud.cropin.in/services/farm/api/farmers/"  # Replace with your GET API URL
PUT_API_URL = "https://cloud.cropin.in/services/farm/api/farmers/sync"  # Replace with your PUT API URL
BEARER_TOKEN = "eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJTbWRPdHp4UWZuT092MjVSUDNseEJNQ2NoWkJYX1pDNG9PU0h3b2hKbHFZIn0.eyJqdGkiOiIxMmFmODAxYy1hZjExLTQxNzctOWZhYi0yZTdlNmU1YTgyMzciLCJleHAiOjE3MzU4OTY3MTYsIm5iZiI6MCwiaWF0IjoxNzM0NjAwNzE2LCJpc3MiOiJodHRwczovL3Nzby5zZy5jcm9waW4uaW4vYXV0aC9yZWFsbXMvcGRsIiwiYXVkIjpbInJlc291cmNlX3NlcnZlciIsImFjY291bnQiXSwic3ViIjoiZjAyODI0ODUtNWU0ZS00ODczLTg0ZTMtMzM4MjcwZGIyYzFjIiwidHlwIjoiQmVhcmVyIiwiYXpwIjoid2ViX2FwcCIsIm5vbmNlIjoiX2NsaWVudF93ZWJfYXBwIiwiYXV0aF90aW1lIjoxNzM0NjAwNzE2LCJzZXNzaW9uX3N0YXRlIjoiNTE2OTNjYzYtODBiMS00YTQ2LWI5YmUtNTZhNWEwYzQyMGFmIiwiYWNyIjoiMSIsImFsbG93ZWQtb3JpZ2lucyI6WyIqIl0sInJlYWxtX2FjY2VzcyI6eyJyb2xlcyI6WyJvZmZsaW5lX2FjY2VzcyIsInVtYV9hdXRob3JpemF0aW9uIl19LCJyZXNvdXJjZV9hY2Nlc3MiOnsicmVzb3VyY2Vfc2VydmVyIjp7InJvbGVzIjpbIlJPTEVfQURNSU5fMTM1MSJdfSwiYWNjb3VudCI6eyJyb2xlcyI6WyJtYW5hZ2UtYWNjb3VudCIsIm1hbmFnZS1hY2NvdW50LWxpbmtzIiwidmlldy1wcm9maWxlIl19fSwic2NvcGUiOiJvcGVuaWQgb2ZmbGluZV9hY2Nlc3MgYWRkcmVzcyBtaWNyb3Byb2ZpbGUtand0IHBob25lIHByb2ZpbGUgZW1haWwiLCJ1c2VyX3JvbGUiOlsiUk9MRV9BRE1JTl8xMzUxIl0sInVwbiI6Ijc2OTg2MTU0MTAiLCJhZGRyZXNzIjp7fSwiZW1haWxfdmVyaWZpZWQiOmZhbHNlLCJncm91cHMiOlsib2ZmbGluZV9hY2Nlc3MiLCJ1bWFfYXV0aG9yaXphdGlvbiJdLCJwcmVmZXJyZWRfdXNlcm5hbWUiOiI3Njk4NjE1NDEwIiwidGVuYW50IjoicGRsIiwiZW1haWwiOiJiYXJhaXlha2FsdWJoYWk0MTBAZ21haWwuY29tIn0.HZ47gwdipEI97zpZ9qcszOID9J2MgSxq5sd94Y8hwfdq2oaUWJKRHCZKh3RXYz1Yb31EH4uTrdFSLyLSX2-jF67YEWdE1Bdd5xUR4rPPHecgQeOzk9tu2k0xxvHMbd54EHM2RLcVlXF2m6pIsPUoRwpzrUtLvdc1E8DaS5D4TVHiwPsxze5MmcRmjGGGz26IKRkgqygIclxLCvQFI-lJW3MDADDg1-QrfnBAKwwpaGtTdGlrBgFGhHFNuvZ_hyPZauc8J69IRChXR1XlUjDdGmejnAncZ2lt4NTnhlEDcnXcwX8VHPI2uLjCHyKi_3kUZFpmlVwnvbR6Jab_CsProA"  # Replace with your actual bearer token

HEADERS = {
    "Authorization": f"Bearer {BEARER_TOKEN}",
    "Content-Type": "application/json"
}

# Load Excel file
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb.active

# Add a status column header if it doesn't exist
if sheet.cell(1, sheet.max_column).value != "Status":
    sheet.cell(1, sheet.max_column + 1, "Status")
status_col = sheet.max_column

# Loop through rows (starting from the second row)
for row in range(2, sheet.max_row + 1):
    farmer_id = sheet.cell(row=row, column=1).value  # Assuming Farmer ID is in column A
    village_name = sheet.cell(row=row, column=4).value  # Assuming Village Name is in column D

    if not farmer_id or not village_name:
        sheet.cell(row=row, column=status_col, value="Skipped: Missing Data")
        continue

    try:
        # GET API call
        get_response = requests.get(f"{GET_API_URL}{farmer_id}", headers=HEADERS)
        get_response.raise_for_status()
        farmer_data = get_response.json()
        print(f"Row = {row}")
        print(f"Getting farmer details for : {farmer_id}")

        # Update the address section with the new village name
        if "address" in farmer_data and isinstance(farmer_data["address"], dict):
            farmer_data["address"]["sublocalityLevel2"] = village_name
            print(f"Village modified for : {farmer_id}")
        else:
            sheet.cell(row=row, column=status_col, value="Failed: No address data")
            continue

        # Wait before making the PUT API call
        time.sleep(1)

        # PUT API call
        put_response = requests.put(PUT_API_URL, headers=HEADERS, json=farmer_data)
        put_response.raise_for_status()

        # Log success in Excel
        sheet.cell(row=row, column=status_col, value="Success")
        print(f"Village updated Successfully for : {farmer_id}")
    except requests.exceptions.RequestException as e:
        # Log failure in Excel
        sheet.cell(row=row, column=status_col, value=f"Failed: {str(e)}")
        print(f"Village update is failed for : {farmer_id}")

    # Wait before processing the next row
    time.sleep(1)

# Save the updated Excel file
output_file = r"C:\Users\rajasekhar.palleti\Downloads\Pidilite_failed_records_updated.xlsx"
wb.save(output_file)
print(f"Excel file updated and saved as {output_file}")
