import requests
import json
import openpyxl
import time

# API URLs
getCAurl = "https://v2uat-gcp.cropin.co.in/uat1/services/farm/api/croppable-areas"
putCAurl = "https://v2uat-gcp.cropin.co.in/uat1/services/farm/api/croppable-areas/area-audit"

# Excel File Path
excel_file_path = r"C:\Users\rajasekhar.palleti\Downloads\areaauditAPI.xlsx"

# API Token
token = "eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJaS08wSFZ2OGlVVmxzQTl6THFBUjhEOWVMc3NYNlVYTERWRkUzdkJ1N0lJIn0.eyJqdGkiOiIzNjc5Mzc5Yy1mZDQzLTQyYzgtYjVmMi1hYjA0MWI5M2NhODEiLCJleHAiOjE3NDE0Mjk2OTcsIm5iZiI6MCwiaWF0IjoxNzQxMjU2ODk3LCJpc3MiOiJodHRwczovL3Yyc3NvLXVhdC1nY3AuY3JvcGluLmNvLmluL2F1dGgvcmVhbG1zL3VhdHpvbmUxIiwiYXVkIjpbInJlc291cmNlX3NlcnZlciIsImFjY291bnQiXSwic3ViIjoiYmYwMmYzM2ItMzZlOC00YmNjLWI2NDUtNmQ4OGM4NmU1N2M5IiwidHlwIjoiQmVhcmVyIiwiYXpwIjoid2ViX2FwcCIsIm5vbmNlIjoiX2NsaWVudF93ZWJfYXBwIiwiYXV0aF90aW1lIjoxNzQxMjU2ODk3LCJzZXNzaW9uX3N0YXRlIjoiYTc2N2QxZWYtZGJjMi00ZjQ1LWJhMWQtNWU2MTI1N2Q4YjQ5IiwiYWNyIjoiMSIsImFsbG93ZWQtb3JpZ2lucyI6WyIqIl0sInJlYWxtX2FjY2VzcyI6eyJyb2xlcyI6WyJvZmZsaW5lX2FjY2VzcyIsInVtYV9hdXRob3JpemF0aW9uIl19LCJyZXNvdXJjZV9hY2Nlc3MiOnsicmVzb3VyY2Vfc2VydmVyIjp7InJvbGVzIjpbIlJPTEVfQURNSU5fMTM1MSJdfSwiYWNjb3VudCI6eyJyb2xlcyI6WyJtYW5hZ2UtYWNjb3VudCIsIm1hbmFnZS1hY2NvdW50LWxpbmtzIiwidmlldy1wcm9maWxlIl19fSwic2NvcGUiOiJvcGVuaWQgbWljcm9wcm9maWxlLWp3dCBwaG9uZSBwcm9maWxlIG9mZmxpbmVfYWNjZXNzIGVtYWlsIGFkZHJlc3MiLCJ1c2VyX3JvbGUiOlsiUk9MRV9BRE1JTl8xMzUxIl0sInVwbiI6Ijk3Mzk3NDcyMTciLCJhZGRyZXNzIjp7fSwiZW1haWxfdmVyaWZpZWQiOmZhbHNlLCJncm91cHMiOlsib2ZmbGluZV9hY2Nlc3MiLCJ1bWFfYXV0aG9yaXphdGlvbiJdLCJwcmVmZXJyZWRfdXNlcm5hbWUiOiI5NzM5NzQ3MjE3IiwidGVuYW50IjoidWF0em9uZTEiLCJlbWFpbCI6InNhdGhpc2guYW5hbmRAY3JvcGluLmNvbSJ9.TFErI542sE9-8DsRqElTxzv0Ohqtyo7ti7ijiMd0z7Gi0JhMkN8vbM-KFbI3seRIIIO6FHy_JDwYU3pGFQsL0gvtecIAm7cjLk0BrbQITqz3d7elPSJHg7gF7TDpbUNR4pFi-OC-14farD8_VSFGmIoq0eeESWxXbdG8pfpVpEa_F8LoXCOAUfuymDhJ-OV2Pmx0ui_B2LZ_6Tj4OelUG9JucgE2XdqLSpIcclt4noYDNksweUKn833sW1n5Rv68YmtH3-BXldcZyKrCOU-1j83lRF95O9hE4evCXMmjHKkUOq3ulcBc1l8YWFxWcae9wXcrl7hwF8GxZSeCqPe2aQ"

# Headers
request_headers = {
    'Authorization': 'Bearer ' + token,
    'Content-Type': "application/json"
}

# Load Excel File
wk = openpyxl.load_workbook(excel_file_path)
sh = wk.active
rows = sh.max_row
print("Rows : ", rows)

c = 0
for i in range(2, sh.max_row + 1):  # Adjusted loop range
    cell_caName = sh.cell(row=i, column=1)
    cell_caId = sh.cell(row=i, column=2)
    cell_areaAudit = sh.cell(row=i, column=3)
    cell_latitude = sh.cell(row=i, column=4)
    cell_longitude = sh.cell(row=i, column=5)
    cell_coordinates = sh.cell(row=i, column=6)
    cell_SuccessOrFail = sh.cell(row=i, column=7)
    cell_Response = sh.cell(row=i, column=8)

    # Validate and Parse cell_coordinates
    if not cell_coordinates.value or cell_coordinates.value.strip() == "":
        print(f"Row {i} - Skipping due to empty coordinates")
        continue

    try:
        str_to_json = json.loads(cell_coordinates.value)
    except json.JSONDecodeError as e:
        print(f"Row {i} - JSONDecodeError: {e}. Skipping this row.")
        continue

    # GET Request to fetch existing data
    resp = requests.get(f"{getCAurl}/{cell_caId.value}", headers=request_headers)

    if resp.status_code != 200:
        print(f"Row {i} - Failed to fetch CA data. Status: {resp.status_code}")
        cell_SuccessOrFail.value = "Failed"
        cell_Response.value = resp.text
        continue

    s_to_j = json.loads(resp.text)

    # Update fields
    s_to_j["cropAudited"] = 'true'  # for Area Audit
    # s_to_j["cropAudited"] = 'false'  # Uncomment for Geotag
    s_to_j["latitude"] = float(cell_latitude.value)
    s_to_j["longitude"] = float(cell_longitude.value)
    s_to_j["auditedArea"] = {"count": float(cell_areaAudit.value)}  # Comment for Geotag
    s_to_j["usableArea"] = {"count": float(cell_areaAudit.value)}  # Comment for Geotag
    s_to_j["areaAudit"] = {
        "geoInfo": str_to_json,
        "latitude": float(cell_latitude.value),
        "longitude": float(cell_longitude.value),
        "channel": 'mobile'
    }

    # PUT Request to update data
    p_response = requests.put(putCAurl, json=s_to_j, headers=request_headers)
    c += 1

    print(f"{c}: {cell_caId.value} - Updated with status {p_response.status_code}")

    # Handle API response
    if p_response.status_code != 200:
        cell_SuccessOrFail.value = "Failed"
        cell_Response.value = p_response.text
    else:
        cell_SuccessOrFail.value = "Success"

    # Save Excel File After Each Iteration
    wk.save(excel_file_path)

    # Delay to prevent server overload
    time.sleep(1)

# Save Final Changes
wk.save(excel_file_path)
print("Process completed!")
