# Author  : Rajasekhar Palleti
# Purpose : Remove FARMER tags from farmers using Excel input.
#
# Description:
# ------------
# This script reads Farmer IDs and Tag identifiers (IDs or Names) from an Excel file,
# fetches farmer details using the Cropin Farmer API, removes matching tags, and
# updates the farmer using PUT API.
#
# Key Features:
# -------------
# - Supports Tag IDs, Tag Names, or mixed input
# - Handles leading/trailing spaces and multiple spaces in tag names
# - Uses Master Tag API to resolve tag names to IDs
# - Skips updates safely when no matching tags are found
# - Saves results to a NEW output Excel file
# - Input Excel file is never modified
#
# Excel Input Format:
# -------------------
# Column A : Farmer ID
# Column B : Tag IDs or Tag Names (comma separated)

import json
import requests
import openpyxl
import time
import re
from GetAuthtoken import get_access_token


# ------------------------------------------------------------
# Normalize tag names (handles spaces & case)
# ------------------------------------------------------------
def normalize_tag_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


# ------------------------------------------------------------
# Fetch FARMER tags (Name → ID mapping)
# ------------------------------------------------------------
def fetch_farmer_tag_map(token):
    url = "https://cloud.cropin.in/services/master/api/filter?type=FARMER&size=10000"
    headers = {"Authorization": f"Bearer {token}"}

    response = requests.get(url, headers=headers)
    response.raise_for_status()

    tag_map = {}
    for tag in response.json():
        if "id" in tag and "name" in tag:
            tag_map[normalize_tag_name(tag["name"])] = tag["id"]

    print(f"✅ Loaded {len(tag_map)} FARMER tags")
    return tag_map


# ------------------------------------------------------------
# Resolve Excel input → Tag IDs
# ------------------------------------------------------------
def resolve_tag_ids(raw_tokens, tag_name_map):
    resolved_ids = []
    unresolved = []

    for token in raw_tokens:
        token = token.strip()

        if token.isdigit():
            resolved_ids.append(int(token))
        else:
            normalized = normalize_tag_name(token)
            if normalized in tag_name_map:
                resolved_ids.append(tag_name_map[normalized])
            else:
                unresolved.append(token)

    return resolved_ids, unresolved


# ------------------------------------------------------------
# Remove tag IDs from farmer response
# ------------------------------------------------------------
def remove_tag_ids(farmer_data, tag_ids_to_remove):
    tags = farmer_data.get("data", {}).get("tags", [])

    if not isinstance(tags, list):
        return farmer_data, set()

    original = set(tags)
    remove_set = set(tag_ids_to_remove)

    updated_tags = [t for t in tags if t not in remove_set]
    removed = original - set(updated_tags)

    farmer_data["data"]["tags"] = updated_tags
    return farmer_data, removed


# ------------------------------------------------------------
# Main processor (matches your existing scripts)
# ------------------------------------------------------------
def process_farmers(api_url, token, input_excel, output_excel, sheet_name):
    wb = openpyxl.load_workbook(input_excel)
    sheet = wb[sheet_name]

    headers = {"Authorization": f"Bearer {token}"}

    # Ensure columns exist
    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    next_col = sheet.max_column + 1

    if "Status" not in headers_row:
        sheet.cell(1, next_col, "Status")
        next_col += 1

    if "Failure Reason" not in headers_row:
        sheet.cell(1, next_col, "Failure Reason")

    headers_row = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    status_col = headers_row.index("Status") + 1
    reason_col = headers_row.index("Failure Reason") + 1

    tag_name_map = fetch_farmer_tag_map(token)

    # --------------------------------------------------------
    # Iterate rows
    # --------------------------------------------------------
    for row in range(2, sheet.max_row + 1):
        farmer_id = sheet.cell(row, 1).value
        raw_tags = sheet.cell(row, 2).value

        if not farmer_id or not raw_tags:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "Missing Farmer ID or Tags")
            continue

        raw_tokens = [t for t in str(raw_tags).split(",") if t.strip()]
        tag_ids, _ = resolve_tag_ids(raw_tokens, tag_name_map)

        if not tag_ids:
            sheet.cell(row, status_col, "Skipped")
            sheet.cell(row, reason_col, "No valid tags found")
            continue

        try:
            get_resp = requests.get(f"{api_url}/{farmer_id}", headers=headers)
            get_resp.raise_for_status()
            farmer_data = get_resp.json()

            existing_tags = farmer_data.get("data", {}).get("tags", [])

            if not set(existing_tags) & set(tag_ids):
                sheet.cell(row, status_col, "Skipped")
                sheet.cell(row, reason_col, "No tags found to remove")
                continue

            updated_farmer, removed_ids = remove_tag_ids(farmer_data, tag_ids)

            multipart_data = {
                "dto": (None, json.dumps(updated_farmer), "application/json")
            }

            time.sleep(0.5)

            put_resp = requests.put(api_url, headers=headers, files=multipart_data)
            put_resp.raise_for_status()

            sheet.cell(row, status_col, "Success")
            sheet.cell(
                row,
                reason_col,
                f"Removed Tag IDs: {', '.join(map(str, removed_ids))}"
            )

        except requests.exceptions.RequestException as e:
            sheet.cell(row, status_col, "Failed")
            sheet.cell(row, reason_col, str(e))

        time.sleep(0.5)

    # ✅ Save to OUTPUT file (same pattern as your scripts)
    wb.save(output_excel)
    print(f"📁 Output saved to: {output_excel}")


# ------------------------------------------------------------
# Entry point
# ------------------------------------------------------------
if __name__ == "__main__":
    input_excel = r"C:\Users\rajasekhar.palleti\Downloads\Farmer_Tags.xlsx"
    output_excel = r"C:\Users\rajasekhar.palleti\Downloads\Farmer_Tags_output.xlsx"
    sheet_name = "Sheet1"
    api_url = "https://cloud.cropin.in/services/farm/api/farmers"

    print("🔐 Fetching access token...")
    token = get_access_token("asp", "9649964096", "123456", "prod1")

    if token:
        process_farmers(api_url, token, input_excel, output_excel, sheet_name)
    else:
        print("❌ Token generation failed.")
