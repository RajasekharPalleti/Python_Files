# Author: Rajasekhar Palleti

import json
import requests
import openpyxl
import time
from GetAuthtoken import get_access_token

def post_data_to_api(api_url, token, input_excel, output_excel):
    # Set up the headers using the retrieved token
    headers = {
        "Authorization": f"Bearer {token}"
    }

    # Load the Excel workbook
    wb = openpyxl.load_workbook(input_excel)
    sheet = wb.active

    # Check and add a "Status" column header if it doesn't exist
    if sheet.cell(1, sheet.max_column).value != "Status":
        sheet.cell(1, sheet.max_column + 1, "Status")
    status_col = sheet.max_column

    # Process each row starting from the second row
    for row in range(2, sheet.max_row + 1):
        asset_id = sheet.cell(row=row, column=1).value   # Assuming Asset ID is in column A
        addlattr_name = sheet.cell(row=row, column=3).value  # Assuming Additional attribute Name is in column C

        if not asset_id or not addlattr_name:
            sheet.cell(row=row, column=status_col, value="Skipped: Missing Data")
            continue

        try:
            # GET API call to retrieve asset data
            get_response = requests.get(f"{api_url}/{asset_id}", headers=headers)
            get_response.raise_for_status()
            asset_data = get_response.json()
            print(f"Row = {row}")
            print(f"Getting asset details for: {asset_id}")

            # Update the additional attribute section with the new additional attribute name
            if "data" in asset_data and isinstance(asset_data["data"], dict):
                if "stateRegistrationinscrioEstadual" not in asset_data["data"] or asset_data["data"]["stateRegistrationinscrioEstadual"] is None:
                    asset_data["data"]["stateRegistrationinscrioEstadual"] = ""
                asset_data["data"]["stateRegistrationinscrioEstadual"] = addlattr_name
                print(f"Additional attribute modified for: {asset_id}")
            else:
                sheet.cell(row=row, column=status_col, value="Failed: No additional attribute data")
                continue

            # Wait before making the PUT API call
            time.sleep(0.2)
            # Convert the payload to a multipart format
            multipart_data = {
                "dto": (None, json.dumps(asset_data), "application/json")  # Convert the entire payload to JSON string
            }

            # PUT API call to update the asset data
            put_response = requests.put(api_url, headers=headers, files=multipart_data)
            put_response.raise_for_status()

            # Log success in Excel
            sheet.cell(row=row, column=status_col, value="Success")
            print(f"Additional attribute is updated successfully for: {asset_id}")
        except requests.exceptions.RequestException as e:
            # Log failure in Excel and print the error
            sheet.cell(row=row, column=status_col, value=f"Failed: {str(e)}")
            print(f"Additional attribute update failed for: {asset_id}")

        # Wait before processing the next row
        time.sleep(0.5)

    # Save the updated Excel file
    wb.save(output_excel)
    print(f"Excel file updated and saved as {output_excel}")

if __name__ == "__main__":
    # Define constants and file paths
    tenant_code = "synlatam"
    input_excel = "C:\\Users\\rajasekhar.palleti\\Downloads\\Upload_Massivo_IE_Soja_Comercial.xlsx"
    output_excel = "C:\\Users\\rajasekhar.palleti\\Downloads\\Upload_Massivo_IE_Soja_Comercial_Updated.xlsx"
    api_url = "https://cloud.cropin.in/services/farm/api/assets"

    # Retrieve access token
    print("Retrieving access token")
    token = get_access_token(tenant_code, "2222255555", "syngenta123")

    if token:
        print("Access token retrieved successfully")
        post_data_to_api(api_url,token, input_excel, output_excel)
    else:
        print("Failed to retrieve access token. Process terminated.")
